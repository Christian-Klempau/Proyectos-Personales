import openpyxl

workbook = openpyxl.load_workbook(filename="Base de Datos Análisis Resultados.xlsx")
hoja = workbook["Resultados"]
print(hoja)

max_col = ""
max_fil = ""

col = 1
fil = 1

elementos = {}
personas = []

while max_col == "":
    if hoja.cell(row=1, column = col).value == None:
        max_col = col-1
        break
    else:
        elementos[hoja.cell(row=1, column = col).value] = col
        col +=1

while max_fil == "":
    if hoja.cell(row=fil, column = 1).value == None:
        max_fil = fil-1
        break
    else:
        fil +=1
     
col_n_eng = elementos["WORK_ENGAGEMENT"]
col_prop_eng = elementos["ENGAGEMENT_PROPORTION"]
col_n_agot = elementos["EXHAUSTION"]
col_prop_agot = elementos["EXHAUSTED_BYNARY"]
col_CAPERF001 = elementos["CAPERF001"]
col_CADI002 = elementos["CADI002"]
col_CAGEN003 = elementos["CAGEN003"]
col_CAAN004 = elementos["CAAN004"]
col_CASED005 = elementos["CASED005"]
set_c1 = set()
set_c2 = set()
set_c3 = set()
set_c4 = set()
set_c5 = set()

class Persona:
    def __init__(self, eng, prop_eng, agot, prop_agot, CAPERF001, CADI002, CAGEN003, CAAN004, CASED005):
        self.eng = eng
        self.prop_eng = prop_eng
        self.agot = agot
        self.prop_agot = prop_agot
        self.CAPERF001 = CAPERF001
        self.CADI002 = CADI002
        self.CAGEN003 = CAGEN003
        self.CAAN004 = CAAN004
        self.CASED005 = CASED005

for i in range(2, max_fil+1):
    eng = hoja.cell(row=i, column=col_n_eng).value
    prop_eng = hoja.cell(row=i, column=col_prop_eng).value
    agot = hoja.cell(row=i, column=col_n_agot).value
    prop_agot = hoja.cell(row=i, column=col_prop_agot).value
    val_C1 = hoja.cell(row=i, column=col_CAPERF001).value
    val_C2 = hoja.cell(row=i, column=col_CADI002).value
    val_C3 = hoja.cell(row=i, column=col_CAGEN003).value
    val_C4 = hoja.cell(row=i, column=col_CAAN004).value
    val_C5 = hoja.cell(row=i, column=col_CASED005).value
    set_c1.add(val_C1)
    set_c2.add(val_C2)
    set_c3.add(val_C3)
    set_c4.add(val_C4)
    set_c5.add(val_C5)
    persona = Persona(eng, prop_eng, agot, prop_agot, val_C1, val_C2, val_C3, val_C4, val_C5)
    personas.append(persona)


# promedio, n total
def promedio(datos):
    tam = len(datos)
    val = sum(datos)
    return (round(val/tam, 2), tam)

# dato = 1, dato = 0
def f_pie_eng(datos):
    c_0 = datos.count(0)
    c_1 = datos.count(1)
    total = c_0 + c_1
    return (round(100*c_0/total, 1), round(100*c_1/total, 1))
# dato = 1, dato = 2, dato = 0
def f_pie_agot(datos):
    c_0 = datos.count(0)
    c_1 = datos.count(1)
    c_2 = datos.count(2)
    total = c_0 + c_1 + c_2
    #print(datos)
    return (int(100*c_0/total), int(100*c_1/total), int(100*c_2/total))


def valores(categoria):
    datos_promedio_eng = []
    datos_promedio_agot = []
    datos_pie_eng = []
    datos_pie_agot = []
    for persona in personas:
        if persona.CADI002 == categoria:
            datos_promedio_eng.append(persona.eng)
            datos_promedio_agot.append(persona.agot)
            datos_pie_eng.append(persona.prop_eng)
            datos_pie_agot.append(persona.prop_agot)

    # tupla promedio, n
    promedio_eng_n = promedio(datos_promedio_eng)
    promedio_agot_n = promedio(datos_promedio_agot)
    # tupla (p0, p1)
    pie_eng = f_pie_eng(datos_pie_eng)
    # tupla (p0, p1, p2)
    pie_agot = f_pie_agot(datos_pie_agot)
    # tupla de tuplas
    # return(categoria, promedio_eng_n, pie_eng, promedio_agot_n, pie_agot)
    s = ""
    s += categoria + ","
    s += str(promedio_eng_n[0]) + ";" + str(promedio_eng_n[1]) + ","
    s += str(pie_eng[0]) + ";" + str(pie_eng[1]) + ","
    s += str(promedio_agot_n[0]) + ";" + str(promedio_agot_n[1]) + ","
    s += str(pie_agot[0]) + ";" + str(pie_agot[1]) + ";" + str(pie_agot[2]) + "\n"
    return s



def crear_excel(categoria, fila):
    file = "resultados"+".xlsx"
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook.active

    vals = valores(categoria)
    print(vals)
    #
    sheet.cell(row=fila+1 , column=1).value = "Val="
    sheet.cell(row=fila+2 , column=1).value = "N="
    sheet.cell(row=fila+1 , column=4).value = 0
    sheet.cell(row=fila+2 , column=4).value = 1
    # 
    sheet.cell(row=fila+1 , column=2).value = vals[0][0]
    sheet.cell(row=fila+2 , column=2).value = vals[0][1]
    #
    sheet.cell(row=fila+5 , column=1).value = "Val="
    sheet.cell(row=fila+6 , column=1).value = "N="
    sheet.cell(row=fila+4 , column=4).value = 0
    sheet.cell(row=fila+5 , column=4).value = 1
    sheet.cell(row=fila+6 , column=4).value = 2
    #
    sheet.cell(row=fila+5 , column=2).value = vals[2][0]
    sheet.cell(row=fila+6 , column=2).value = vals[2][1]
    #
    sheet.cell(row=fila+1 , column=5).value = str(vals[1][0]).replace(",", ".")+"%"
    sheet.cell(row=fila+2 , column=5).value = str(vals[1][1]).replace(",", ".")+"%"
    #
    sheet.cell(row=fila+4 , column=5).value = str(vals[3][0]).replace(",", ".")+"%"
    sheet.cell(row=fila+5 , column=5).value = str(vals[3][1]).replace(",", ".")+"%"
    sheet.cell(row=fila+6 , column=5).value = str(vals[3][2]).replace(",", ".")+"%"
    #
    sheet.cell(row=fila, column=1).value = categoria
    workbook.save(filename=file)

fila = 1
filename = "resultados.xlsx"
workbook = openpyxl.Workbook()
workbook.save(filename=filename)

datos = open("datos.csv", "w", encoding="UTF-8")

for categoria in set_c2:
    vals = valores(categoria)
    #crear_excel(categoria, fila)
    #fila += 8
    print(vals)
    datos.write(vals)
   

datos.close()